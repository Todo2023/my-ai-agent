"""指定した顧客名の請求書PDFを、Excelの請求データから作る。

このファイルにはAI(LLM)は一切登場しない。
「表を読む→計算する→PDFに書く」という決まった処理だけを行う。

実際の請求書フォーマット（発行者情報・請求書番号・税率区分・支払期日）に
合わせている。発行者情報は固定（合同会社To do）で、請求書ごとに変わらない。
"""

import datetime as dt
import json
import sys
from pathlib import Path

import openpyxl
from fpdf import FPDF

FONT_PATH = "/usr/share/fonts/opentype/ipafont-gothic/ipag.ttf"

ISSUER = {
    "name": "合同会社To do",
    "postal_code": "〒004-0022",
    "address": "札幌市厚別区厚別南1-4-3",
    "address2": "Asterope厚別南101",
    "tel": "070-9136-4879",
    "registration_number": "T4011603004093",
    "bank": "GMOあおぞらネット銀行　法人営業部",
    "bank_account": "口座番号：1785630　名義：ド）トゥードゥー",
}

# 顧客マスタに無い顧客に使うデフォルト値
DEFAULT_CONTACT = "辰巳"
DEFAULT_PHRASE = "下記の通り、ご請求申し上げます。"
DEFAULT_HONORIFIC = "御中"

COUNTER_FILE = Path("invoice_counter.json")


def next_invoice_number(issue_date: dt.date) -> str:
    """「YYYYMM-連番」形式の請求書番号を発行する（月ごとに1から連番）。"""
    key = issue_date.strftime("%Y%m")
    counts = json.loads(COUNTER_FILE.read_text()) if COUNTER_FILE.exists() else {}
    counts[key] = counts.get(key, 0) + 1
    COUNTER_FILE.write_text(json.dumps(counts, ensure_ascii=False, indent=2))
    return f"{key}-{counts[key]:03d}"


def compute_due_date(issue_date: dt.date) -> dt.date:
    """支払期日 = 発行月の翌月末日。"""
    if issue_date.month == 12:
        next_month_first = dt.date(issue_date.year + 1, 1, 1)
    else:
        next_month_first = dt.date(issue_date.year, issue_date.month + 1, 1)
    if next_month_first.month == 12:
        month_after_first = dt.date(next_month_first.year + 1, 1, 1)
    else:
        month_after_first = dt.date(next_month_first.year, next_month_first.month + 1, 1)
    return month_after_first - dt.timedelta(days=1)


def list_customer_names(excel_path: str) -> list[str]:
    """Excelに登録されている顧客名の一覧を、重複なしで返す。"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["請求データ"]

    names = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if name not in names:
            names.append(name)
    return names


def load_customer_master(excel_path: str, customer_name: str) -> dict:
    """「顧客マスタ」シートから、担当者・請求文言・敬称を取得する。

    シートが無い、またはその顧客の行が無い場合はデフォルト値を使う。
    列構成: 顧客名 / 担当者 / 請求文言 / 敬称（任意。個人宛なら「様」など。空欄は「御中」）
    """
    defaults = {"contact": DEFAULT_CONTACT, "phrase": DEFAULT_PHRASE, "honorific": DEFAULT_HONORIFIC}
    wb = openpyxl.load_workbook(excel_path)
    if "顧客マスタ" not in wb.sheetnames:
        return defaults

    ws = wb["顧客マスタ"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, contact, phrase = row[0], row[1], row[2]
        honorific = row[3] if len(row) > 3 else None
        if name == customer_name:
            return {
                "contact": contact or DEFAULT_CONTACT,
                "phrase": phrase or DEFAULT_PHRASE,
                "honorific": honorific or DEFAULT_HONORIFIC,
            }
    return defaults


def load_items_for_customer(excel_path: str, customer_name: str) -> dict:
    """Excelから、指定した顧客の行を取り出し、明細と件名・納品日をまとめて返す。

    列構成: 顧客名 / 日付 / 品目 / 数量 / 単価 / 税率 / 件名 / 納品日
    件名・納品日は同じ顧客の行で共通の値を想定し、最初に見つかった値を使う。
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["請求データ"]

    line_items = []
    matter = None
    delivery_date = None

    remarks = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        name, date_, item, quantity, unit_price, tax_rate, row_matter, row_delivery_date = row[:8]
        row_remarks = row[8] if len(row) > 8 else None
        if name != customer_name:
            continue
        quantity = quantity or 0
        unit_price = unit_price or 0
        line_items.append(
            {
                "date": date_,
                "item": item,
                "quantity": quantity,
                "unit_price": unit_price,
                "tax_rate": tax_rate if tax_rate else 10,
                "subtotal": quantity * unit_price,
            }
        )
        if matter is None and row_matter:
            matter = row_matter
        if delivery_date is None and row_delivery_date:
            delivery_date = row_delivery_date
        if remarks is None and row_remarks:
            remarks = row_remarks

    return {"line_items": line_items, "matter": matter, "delivery_date": delivery_date, "remarks": remarks}


def _format_date(value) -> str:
    if isinstance(value, dt.datetime):
        value = value.date()
    if isinstance(value, dt.date):
        return f"{value.month}月{value.day}日"
    return str(value)


def create_invoice_pdf(
    customer_name: str, excel_path: str, output_path: str, issue_date: dt.date | None = None
) -> None:
    """指定した顧客の請求書PDFを作る。issue_date省略時は本日日付。"""
    data = load_items_for_customer(excel_path, customer_name)
    line_items = data["line_items"]

    if not line_items:
        raise ValueError(f"「{customer_name}」のデータが見つかりませんでした")

    master = load_customer_master(excel_path, customer_name)

    subtotal_10 = sum(i["subtotal"] for i in line_items if i["tax_rate"] == 10)
    subtotal_8 = sum(i["subtotal"] for i in line_items if i["tax_rate"] == 8)
    tax_10 = round(subtotal_10 * 0.10)
    tax_8 = round(subtotal_8 * 0.08)
    subtotal = subtotal_10 + subtotal_8
    total = subtotal + tax_10 + tax_8

    issue_date = issue_date or dt.date.today()
    due_date = compute_due_date(issue_date)
    invoice_number = next_invoice_number(issue_date)

    pdf = FPDF(format="A4")
    pdf.set_margins(15, 15, 15)
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.add_font("IPAGothic", "", FONT_PATH)
    page_w = pdf.w - pdf.l_margin - pdf.r_margin  # 180mm

    # 発行日・No（右上）
    pdf.set_xy(15, 15)
    pdf.set_font("IPAGothic", size=9)
    pdf.cell(page_w, 5, f"発行日：{issue_date.year}年{issue_date.month}月{issue_date.day}日", align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.set_x(15)
    pdf.cell(page_w, 5, f"No：{invoice_number}", align="R", new_x="LMARGIN", new_y="NEXT")

    pdf.ln(6)

    # タイトル
    pdf.set_font("IPAGothic", size=22)
    pdf.set_x(15)
    pdf.cell(page_w, 12, "請求書", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    y0 = pdf.get_y()

    # 宛先（左、下線＋御中）
    pdf.set_xy(15, y0)
    pdf.set_font("IPAGothic", size=12)
    pdf.cell(95, 8, f"{customer_name}", border="B")
    pdf.cell(15, 8, master["honorific"])

    # 発行者情報（右側に重ねて描く固定ブロック）
    right_x = 120
    right_w = 75
    ry = y0 - 2
    issuer_lines = [
        (ISSUER["name"], 10),
        (f"{ISSUER['postal_code']} {ISSUER['address']}", 9),
        (ISSUER["address2"], 9),
        (f"TEL：{ISSUER['tel']}", 9),
        (f"担当：{master['contact']}", 9),
        (f"登録番号：{ISSUER['registration_number']}", 9),
    ]
    for text, size in issuer_lines:
        pdf.set_xy(right_x, ry)
        pdf.set_font("IPAGothic", size=size)
        pdf.cell(right_w, 5.5, text, align="R")
        ry += 5.5

    pdf.set_xy(15, max(y0 + 12, ry) + 4)

    # 件名・納品日（ラベル＋下線）
    pdf.set_font("IPAGothic", size=10)
    pdf.cell(20, 7, "件　名：")
    pdf.cell(130, 7, data["matter"] or "", border="B", new_x="LMARGIN", new_y="NEXT")
    pdf.set_x(15)
    pdf.ln(2)
    pdf.set_x(15)
    pdf.cell(20, 7, "納品日：")
    pdf.cell(60, 7, _format_date(data["delivery_date"]) if data["delivery_date"] else "", border="B")
    pdf.ln(12)

    pdf.set_x(15)
    pdf.set_font("IPAGothic", size=9)
    pdf.cell(page_w, 6, master["phrase"], new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    # 金額（大きく、下線つき）
    pdf.set_x(15)
    pdf.set_font("IPAGothic", size=11)
    pdf.cell(20, 10, "金額")
    pdf.set_font("IPAGothic", size=18)
    pdf.cell(70, 10, f"¥{total:,}", border="B")
    pdf.ln(16)

    # 明細テーブル
    col_widths = [25, 85, 20, 25, 25]
    headers = ["日付", "項目", "数量", "単価", "金額"]
    pdf.set_x(15)
    pdf.set_font("IPAGothic", size=9)
    pdf.set_fill_color(50, 50, 50)
    pdf.set_text_color(255, 255, 255)
    for w, h in zip(col_widths, headers):
        pdf.cell(w, 8, h, border=1, align="C", fill=True)
    pdf.ln(8)
    pdf.set_text_color(0, 0, 0)

    shaded = False
    for rate in (10, 8):
        group = [i for i in line_items if i["tax_rate"] == rate]
        if not group:
            continue
        pdf.set_x(15)
        pdf.set_fill_color(210, 225, 240)
        pdf.cell(sum(col_widths), 7, f"●税率{rate}%項目", border=1, fill=True, new_x="LMARGIN", new_y="NEXT")
        for i in group:
            pdf.set_x(15)
            shaded = not shaded
            pdf.set_fill_color(235, 242, 250) if shaded else pdf.set_fill_color(255, 255, 255)
            values = [
                _format_date(i["date"]),
                str(i["item"]),
                str(i["quantity"]) if i["quantity"] else "",
                f"{i['unit_price']:,}" if i["unit_price"] else "",
                f"{i['subtotal']:,}" if i["subtotal"] else "",
            ]
            for w, v in zip(col_widths, values):
                pdf.cell(w, 7, v, border=1, fill=True)
            pdf.ln(7)

    pdf.ln(4)

    # 税率別の集計
    label_w, val_w = 140, 40
    pdf.set_font("IPAGothic", size=10)
    for label, value in [
        ("小計", subtotal),
        ("消費税(10%対象)", tax_10),
        ("消費税(8%対象)", tax_8),
    ]:
        pdf.set_x(15)
        pdf.cell(label_w, 7, label, align="R")
        pdf.cell(val_w, 7, f"{value:,}", align="R", new_x="LMARGIN", new_y="NEXT")

    pdf.set_x(15)
    pdf.set_font("IPAGothic", size=12)
    pdf.cell(label_w, 9, "税込合計", align="R")
    pdf.cell(val_w, 9, f"¥{total:,}", align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(10)

    if data["remarks"]:
        pdf.set_x(15)
        pdf.set_font("IPAGothic", size=9)
        pdf.cell(page_w, 6, f"備考　{data['remarks']}", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

    # 振込先・支払期日
    pdf.set_x(15)
    pdf.set_font("IPAGothic", size=9)
    pdf.cell(page_w, 6, f"振込先　{ISSUER['bank']}", new_x="LMARGIN", new_y="NEXT")
    pdf.set_x(15)
    pdf.cell(page_w, 6, f"　　　　{ISSUER['bank_account']}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)
    pdf.set_x(15)
    pdf.cell(page_w, 6, f"支払期日　{due_date.year}年{due_date.month}月{due_date.day}日", new_x="LMARGIN", new_y="NEXT")

    pdf.output(output_path)
    print(f"{output_path} を作成しました（No:{invoice_number}, 税込合計: {total:,}円）")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("使い方: uv run invoice.py 顧客名")
        sys.exit(1)

    customer = sys.argv[1]
    create_invoice_pdf(customer, "sample_data.xlsx", f"invoice_{customer}.pdf")
