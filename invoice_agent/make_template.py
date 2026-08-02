"""実データ入力用の請求書テンプレート(Excel)を作る使い捨てスクリプト。

実行すると invoice_template.xlsx が作られる。
「請求データ」シートに入力する。「記入例」の行は削除してから使う。
"""

import datetime as dt

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "請求データ"

headers = ["顧客名", "日付", "品目", "数量", "単価", "税率", "件名", "納品日"]
ws.append(headers)

header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col)
    cell.font = Font(bold=True)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

notes = {
    "A1": "請求書エージェントが顧客を検索するキー。表記ゆれ（全角半角・「様」の有無など）に注意",
    "B1": "その作業・納品を行った日付（例: 2026-07-16）",
    "C1": "請求書の明細に載る項目名",
    "D1": "数量。金額固定の項目は1でよい",
    "E1": "単価（円）。同じ件名の中で金額を1行にまとめたい場合、他の行は0でよい",
    "F1": "消費税率。10 か 8 を入力（空欄は10として扱われる）",
    "G1": "請求書の「件名」欄に載る。同じ請求書に含める行はすべて同じ値にする",
    "H1": "請求書の「納品日」欄に載る。同じ請求書に含める行はすべて同じ値にする",
}
for ref, text in notes.items():
    ws[ref].comment = Comment(text, "invoice_agent")

# 記入例（コピーして使う。使う前に削除してよい）
ws.append(["（記入例）A社", dt.date(2026, 7, 16), "MTG", 1, 75000, 10, "営業資料作成 他", dt.date(2026, 7, 27)])
ws.append(["（記入例）A社", dt.date(2026, 7, 22), "MTG（すり合わせ）", 1, 0, 10, "営業資料作成 他", dt.date(2026, 7, 27)])
ws.append(["（記入例）A社", dt.date(2026, 7, 27), "資料納品", 1, 0, 10, "営業資料作成 他", dt.date(2026, 7, 27)])

for col, width in zip(range(1, 9), [16, 12, 28, 8, 10, 8, 24, 12]):
    ws.column_dimensions[chr(64 + col)].width = width

wb.save("invoice_template.xlsx")
print("invoice_template.xlsx を作成しました")
